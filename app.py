"""
2025년 병원 처방 내역 및 청구/삭감 비교 분석 대시보드
Hospital Management Consulting Dashboard - Streamlit App
"""

import streamlit as st
import pandas as pd
import numpy as np
import plotly.express as px
import plotly.graph_objects as go
from plotly.subplots import make_subplots
from openpyxl import load_workbook
import os
import re
import warnings
warnings.filterwarnings("ignore")

# ─────────────────────────────────────────────
# 페이지 설정
# ─────────────────────────────────────────────
st.set_page_config(
    page_title="2025 병원 경영 분석 대시보드",
    page_icon="🏥",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ─────────────────────────────────────────────
# CSS 스타일
# ─────────────────────────────────────────────
st.markdown("""
<style>
    /* 전체 배경 */
    .stApp { background-color: #f0f4f8; }
    
    /* 사이드바 */
    [data-testid="stSidebar"] { background-color: #1a2744; }
    [data-testid="stSidebar"] * { color: #e8edf5 !important; }
    [data-testid="stSidebar"] .stSelectbox label,
    [data-testid="stSidebar"] .stMultiSelect label { color: #a8b8d8 !important; }

    /* KPI 카드 */
    .kpi-card {
        background: white;
        border-radius: 12px;
        padding: 20px 24px;
        box-shadow: 0 2px 8px rgba(0,0,0,0.08);
        border-left: 5px solid #2563eb;
        margin-bottom: 8px;
    }
    .kpi-card.red   { border-left-color: #dc2626; }
    .kpi-card.green { border-left-color: #16a34a; }
    .kpi-card.amber { border-left-color: #d97706; }
    .kpi-card.purple{ border-left-color: #7c3aed; }
    .kpi-label { font-size: 13px; color: #6b7280; font-weight: 500; margin-bottom: 6px; }
    .kpi-value { font-size: 26px; font-weight: 700; color: #111827; }
    .kpi-sub   { font-size: 12px; color: #9ca3af; margin-top: 4px; }

    /* 섹션 헤더 */
    .section-header {
        background: linear-gradient(135deg, #1a2744 0%, #2563eb 100%);
        color: white;
        padding: 12px 20px;
        border-radius: 10px;
        font-size: 18px;
        font-weight: 700;
        margin: 20px 0 12px 0;
    }

    /* 조정사유 박스 */
    .reason-box {
        background: #fef9c3;
        border-left: 4px solid #eab308;
        padding: 10px 14px;
        border-radius: 6px;
        font-size: 13px;
        color: #713f12;
        margin: 4px 0;
    }

    /* 테이블 스타일 */
    .dataframe { font-size: 13px !important; }

    /* 페이지 타이틀 */
    .main-title {
        background: linear-gradient(135deg, #1a2744 0%, #1e40af 50%, #3b82f6 100%);
        color: white;
        padding: 28px 36px;
        border-radius: 14px;
        margin-bottom: 24px;
        box-shadow: 0 4px 20px rgba(37,99,235,0.25);
    }
    .main-title h1 { font-size: 28px; font-weight: 800; margin: 0; }
    .main-title p  { font-size: 14px; opacity: 0.8; margin: 6px 0 0 0; }

    /* 탭 스타일 */
    .stTabs [data-baseweb="tab"] {
        font-size: 14px; font-weight: 600; padding: 10px 20px;
    }
    .stTabs [aria-selected="true"] {
        background-color: #2563eb; color: white; border-radius: 8px 8px 0 0;
    }
</style>
""", unsafe_allow_html=True)


# ─────────────────────────────────────────────
# 데이터 로드 함수
# ─────────────────────────────────────────────

MONTHS = ['1월','2월','3월','4월','5월','6월','7월','8월','9월','10월','11월','12월']
MONTH_NUM = {m: i+1 for i, m in enumerate(MONTHS)}

def parse_amount(val):
    """문자열 금액 → 숫자 변환"""
    if val is None or val == '' or (isinstance(val, float) and np.isnan(val)):
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)
    return float(str(val).replace(',', '').strip() or 0)

@st.cache_data
def load_billing_data(filepath):
    """청구금액_및_삭감내역.xlsx 로드"""
    wb = load_workbook(filepath, read_only=True)
    result = {}

    for sheet in wb.sheetnames:
        ws = wb[sheet]
        rows = list(ws.iter_rows(values_only=True))

        # 외래 / 입원 블록을 구분
        for section_label, data_key in [('외래', 'outpatient'), ('입원', 'inpatient')]:
            records = []
            in_block = False
            for row in rows:
                text = str(row[1]) if row[1] else ''
                if section_label in text and '청구금액' in text:
                    in_block = True
                    continue
                if in_block:
                    if row[1] in ('월별',):
                        continue
                    if row[1] in MONTHS:
                        records.append({
                            '월': row[1],
                            '요양급여총액': parse_amount(row[2]),
                            '건수': parse_amount(row[3]),
                            '청구액': parse_amount(row[4]),
                            '조정액': parse_amount(row[5]),
                            '심사결정액': parse_amount(row[7]),
                            '주요조정사유': str(row[8]) if row[8] else '',
                        })
                    elif row[1] == '합계':
                        in_block = False

            if records:
                df = pd.DataFrame(records)
                df['월번호'] = df['월'].map(MONTH_NUM)
                df['조정비율'] = df.apply(
                    lambda r: (r['조정액'] / r['청구액'] * 100) if r['청구액'] > 0 else 0, axis=1
                )
                result[(sheet, data_key)] = df

    return result


@st.cache_data
def load_prescription_data(filepath, kind='outpatient'):
    """처방 내역 xlsx 로드 → 월별/항목별/의사별 집계"""
    wb = load_workbook(filepath, read_only=True)
    sheet_suffix = '외래' if kind == 'outpatient' else '입원'

    monthly_summary = []   # 월별 총 집계
    monthly_detail  = []   # 월별 항목별 집계
    doctor_monthly  = {}   # 의사별 월별 집계

    for sh in wb.sheetnames:
        if sheet_suffix not in sh:
            continue
        # 월 추출
        month_match = re.search(r'(\d+)월', sh)
        if not month_match:
            continue
        month_label = f"{month_match.group(1)}월"

        ws = wb[sh]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            continue

        # 헤더 행(row 0): 의사 정보
        header_row = rows[0]
        # 열 구조: [항목, 소계-총투, 소계-금액, 의사1-총투, 의사1-금액, ...]
        doctors = []
        col_idx = 1  # 소계
        while col_idx < len(header_row):
            name = header_row[col_idx]
            if name and name != '소계':
                doctors.append((name, col_idx))
            col_idx += 2

        total_qty   = 0.0
        total_amt   = 0.0
        item_records = {}

        for row in rows[2:]:  # 헤더 2줄 skip
            item = row[0]
            if not item or item in ('항목(소분류)', '합계'):
                continue
            qty = parse_amount(row[1])
            amt = parse_amount(row[2])
            total_qty += qty
            total_amt += amt

            item_records[item] = item_records.get(item, {'총투': 0.0, '금액': 0.0})
            item_records[item]['총투'] += qty
            item_records[item]['금액'] += amt

            # 의사별
            for doc_name, col in doctors:
                dq = parse_amount(row[col])     if col < len(row) else 0
                da = parse_amount(row[col+1])   if col+1 < len(row) else 0
                key = (doc_name, month_label)
                if key not in doctor_monthly:
                    doctor_monthly[key] = {'총투': 0.0, '금액': 0.0}
                doctor_monthly[key]['총투'] += dq
                doctor_monthly[key]['금액'] += da

        monthly_summary.append({
            '월': month_label,
            '총투': total_qty,
            '금액': total_amt,
        })
        for item, vals in item_records.items():
            monthly_detail.append({
                '월': month_label,
                '항목': item,
                '총투': vals['총투'],
                '금액': vals['금액'],
            })

    df_summary = pd.DataFrame(monthly_summary)
    df_summary['월번호'] = df_summary['월'].map(MONTH_NUM)
    df_summary = df_summary.sort_values('월번호').reset_index(drop=True)

    df_detail = pd.DataFrame(monthly_detail)
    if not df_detail.empty:
        df_detail['월번호'] = df_detail['월'].map(MONTH_NUM)
        df_detail = df_detail.sort_values(['월번호','항목']).reset_index(drop=True)

    # 의사별 데이터프레임
    doc_records = []
    for (doc, mon), vals in doctor_monthly.items():
        doc_records.append({'의사': doc, '월': mon, '총투': vals['총투'], '금액': vals['금액']})
    df_doctor = pd.DataFrame(doc_records) if doc_records else pd.DataFrame(columns=['의사','월','총투','금액'])
    if not df_doctor.empty:
        df_doctor['월번호'] = df_doctor['월'].map(MONTH_NUM)
        df_doctor = df_doctor.sort_values(['의사','월번호']).reset_index(drop=True)

    return df_summary, df_detail, df_doctor


def fmt_won(val):
    """원화 포맷"""
    if abs(val) >= 1e8:
        return f"{val/1e8:.1f}억원"
    elif abs(val) >= 1e4:
        return f"{val/1e4:,.0f}만원"
    return f"{val:,.0f}원"

def fmt_pct(val):
    return f"{val:.2f}%"

# ─────────────────────────────────────────────
# 사이드바
# ─────────────────────────────────────────────
with st.sidebar:
    st.markdown("## 🏥 병원 경영 분석")
    st.markdown("---")

    st.markdown("### 📁 데이터 파일 설정")
    billing_file  = st.text_input("청구/삭감 파일", value="2025_청구금액_및_삭감내역.xlsx")
    outpat_file   = st.text_input("외래 처방 파일", value="2025처방_외래_.xlsx")
    inpat_file    = st.text_input("입원 처방 파일", value="2025처방_입원_.xlsx")

    st.markdown("---")
    st.markdown("### 📅 분석 기간")
    month_range = st.select_slider(
        "월 범위 선택",
        options=MONTHS,
        value=('1월', '12월')
    )
    sel_months = MONTHS[MONTHS.index(month_range[0]):MONTHS.index(month_range[1])+1]

    st.markdown("---")
    st.markdown("### 🔍 분석 구분")
    view_type = st.radio("구분", ["외래", "입원", "전체 비교"], index=2)

    st.markdown("---")
    st.caption("ⓒ 2025 병원 경영 컨설팅")


# ─────────────────────────────────────────────
# 타이틀
# ─────────────────────────────────────────────
st.markdown("""
<div class="main-title">
  <h1>🏥 2025년 병원 경영 분석 대시보드</h1>
  <p>처방 내역 · 청구금액 · 삭감내역 종합 분석 | 건강보험 · 의료급여</p>
</div>
""", unsafe_allow_html=True)


# ─────────────────────────────────────────────
# 데이터 로드
# ─────────────────────────────────────────────
def convert_xls_to_xlsx(xls_path):
    """xls → xlsx 변환 (xlrd + openpyxl, 윈도우 호환)"""
    from openpyxl import Workbook
    dst_dir = os.path.dirname(xls_path)
    basename = os.path.basename(xls_path)
    xlsx_name = os.path.splitext(basename)[0] + "_converted.xlsx"
    xlsx_path = os.path.join(dst_dir, xlsx_name)
    if os.path.exists(xlsx_path):
        return xlsx_path
    try:
        import xlrd
        xls_wb = xlrd.open_workbook(xls_path)
        new_wb = Workbook()
        new_wb.remove(new_wb.active)
        for sheet_name in xls_wb.sheet_names():
            xls_ws = xls_wb.sheet_by_name(sheet_name)
            new_ws = new_wb.create_sheet(title=sheet_name)
            for row_idx in range(xls_ws.nrows):
                for col_idx in range(xls_ws.ncols):
                    cell = xls_ws.cell(row_idx, col_idx)
                    new_ws.cell(row=row_idx+1, column=col_idx+1, value=cell.value)
        new_wb.save(xlsx_path)
        return xlsx_path
    except Exception as e:
        st.error(f"xls 변환 실패: {e}\n\n'pip install xlrd' 를 실행한 뒤 다시 시도해 주세요.")
        st.stop()

# 파일 경로 자동 탐색
def find_file(name, search_dirs=None):
    if search_dirs is None:
        search_dirs = ['.', '/mnt/user-data/uploads', '/tmp', os.path.expanduser('~')]
    for d in search_dirs:
        p = os.path.join(d, name)
        if os.path.exists(p):
            return p
    return None

with st.spinner("📊 데이터를 불러오는 중입니다..."):
    billing_path = find_file(billing_file)
    outpat_path  = find_file(outpat_file)
    inpat_path   = find_file(inpat_file)

    errors = []
    if not billing_path: errors.append(f"❌ 청구/삭감 파일 없음: {billing_file}")
    if not outpat_path:  errors.append(f"❌ 외래 처방 파일 없음: {outpat_file}")
    if not inpat_path:   errors.append(f"❌ 입원 처방 파일 없음: {inpat_file}")
    if errors:
        for e in errors:
            st.error(e)
        st.stop()

    # xls → xlsx 변환
    if outpat_path.endswith('.xls'):
        outpat_path = convert_xls_to_xlsx(outpat_path)

    billing_data = load_billing_data(billing_path)
    df_out_sum, df_out_det, df_out_doc = load_prescription_data(outpat_path, 'outpatient')
    df_inp_sum, df_inp_det, df_inp_doc = load_prescription_data(inpat_path, 'inpatient')

    # 필터링
    df_out_sum_f = df_out_sum[df_out_sum['월'].isin(sel_months)]
    df_inp_sum_f = df_inp_sum[df_inp_sum['월'].isin(sel_months)]
    df_out_det_f = df_out_det[df_out_det['월'].isin(sel_months)]
    df_inp_det_f = df_inp_det[df_inp_det['월'].isin(sel_months)]

    # 청구 데이터 병합
    def get_billing(sheet, kind):
        key = (sheet, kind)
        df = billing_data.get(key, pd.DataFrame())
        if not df.empty:
            df = df[df['월'].isin(sel_months)].copy()
        return df

    hb_out = get_billing('건강보험', 'outpatient')
    hb_inp = get_billing('건강보험', 'inpatient')
    mg_out = get_billing('의료급여', 'outpatient')
    mg_inp = get_billing('의료급여', 'inpatient')

    # 전체 통합
    def combine_billing(dfs):
        dfs = [d for d in dfs if not d.empty]
        if not dfs: return pd.DataFrame()
        df = pd.concat(dfs).groupby('월').agg(
            요양급여총액=('요양급여총액','sum'),
            건수=('건수','sum'),
            청구액=('청구액','sum'),
            조정액=('조정액','sum'),
            심사결정액=('심사결정액','sum'),
        ).reset_index()
        df['월번호'] = df['월'].map(MONTH_NUM)
        df['조정비율'] = df.apply(lambda r: r['조정액']/r['청구액']*100 if r['청구액']>0 else 0, axis=1)
        return df.sort_values('월번호')

    all_bill_out = combine_billing([hb_out, mg_out])
    all_bill_inp = combine_billing([hb_inp, mg_inp])
    all_bill_all = combine_billing([hb_out, mg_out, hb_inp, mg_inp])

st.success("✅ 데이터 로드 완료")

# ─────────────────────────────────────────────
# KPI 섹션
# ─────────────────────────────────────────────
st.markdown('<div class="section-header">📈 핵심 경영 지표 (KPI)</div>', unsafe_allow_html=True)

def safe_sum(df, col):
    if df.empty or col not in df.columns: return 0
    return df[col].sum()

total_claim     = safe_sum(all_bill_all, '청구액')
total_adj       = safe_sum(all_bill_all, '조정액')
total_decided   = safe_sum(all_bill_all, '심사결정액')
total_cases     = safe_sum(all_bill_all, '건수')
adj_rate        = total_adj / total_claim * 100 if total_claim > 0 else 0

out_rx_total    = df_out_sum_f['금액'].sum() if not df_out_sum_f.empty else 0
inp_rx_total    = df_inp_sum_f['금액'].sum() if not df_inp_sum_f.empty else 0
rx_total        = out_rx_total + inp_rx_total

k1, k2, k3, k4 = st.columns(4)
with k1:
    st.markdown(f"""
    <div class="kpi-card">
        <div class="kpi-label">💰 총 청구금액</div>
        <div class="kpi-value">{fmt_won(total_claim)}</div>
        <div class="kpi-sub">건수: {total_cases:,.0f}건</div>
    </div>""", unsafe_allow_html=True)
with k2:
    st.markdown(f"""
    <div class="kpi-card red">
        <div class="kpi-label">✂️ 총 조정(삭감)액</div>
        <div class="kpi-value">{fmt_won(total_adj)}</div>
        <div class="kpi-sub">조정비율: {fmt_pct(adj_rate)}</div>
    </div>""", unsafe_allow_html=True)
with k3:
    st.markdown(f"""
    <div class="kpi-card green">
        <div class="kpi-label">✅ 심사결정액</div>
        <div class="kpi-value">{fmt_won(total_decided)}</div>
        <div class="kpi-sub">청구 대비: {fmt_pct(total_decided/total_claim*100 if total_claim>0 else 0)}</div>
    </div>""", unsafe_allow_html=True)
with k4:
    st.markdown(f"""
    <div class="kpi-card purple">
        <div class="kpi-label">💊 총 처방금액</div>
        <div class="kpi-value">{fmt_won(rx_total)}</div>
        <div class="kpi-sub">외래 {fmt_won(out_rx_total)} / 입원 {fmt_won(inp_rx_total)}</div>
    </div>""", unsafe_allow_html=True)


# ─────────────────────────────────────────────
# 탭 구성
# ─────────────────────────────────────────────
tab1, tab2, tab3, tab4, tab5, tab6 = st.tabs([
    "📊 월별 청구·삭감 현황",
    "💊 처방 분석",
    "👨‍⚕️ 의사별 처방",
    "🔍 항목별 분석",
    "⚠️ 삭감 사유 분석",
    "📋 원본 데이터"
])


# ─────────────────────────────────────────────
# TAB 1: 월별 청구·삭감 현황
# ─────────────────────────────────────────────
with tab1:
    st.markdown('<div class="section-header">📊 월별 청구·조정·심사결정 현황</div>', unsafe_allow_html=True)

    col_sel = st.columns(3)
    with col_sel[0]:
        bill_type = st.selectbox("보험 구분", ["전체", "건강보험", "의료급여"])
    with col_sel[1]:
        care_type = st.selectbox("진료 구분", ["전체", "외래", "입원"])
    with col_sel[2]:
        chart_kind = st.selectbox("차트 종류", ["막대+선", "누적 막대", "영역"])

    # 데이터 선택
    def select_bill_df(bill_t, care_t):
        if bill_t == "전체":
            if care_t == "전체":   return all_bill_all
            if care_t == "외래":   return all_bill_out
            if care_t == "입원":   return all_bill_inp
        else:
            sheet = "건강보험" if bill_t == "건강보험" else "의료급여"
            if care_t == "전체":
                return combine_billing([get_billing(sheet,'outpatient'), get_billing(sheet,'inpatient')])
            kind = 'outpatient' if care_t == "외래" else 'inpatient'
            return get_billing(sheet, kind)
        return pd.DataFrame()

    bdf = select_bill_df(bill_type, care_type)

    if bdf.empty:
        st.info("해당 조건의 데이터가 없습니다.")
    else:
        # 주요 지표 요약
        m1, m2, m3, m4 = st.columns(4)
        with m1:
            st.metric("청구액 합계", fmt_won(bdf['청구액'].sum()))
        with m2:
            st.metric("조정액 합계", fmt_won(bdf['조정액'].sum()),
                      delta=f"-{fmt_won(bdf['조정액'].sum())}", delta_color="inverse")
        with m3:
            st.metric("평균 조정비율", fmt_pct(bdf['조정비율'].mean()))
        with m4:
            max_adj_row = bdf.loc[bdf['조정비율'].idxmax()]
            st.metric("최고 조정비율 월", f"{max_adj_row['월']} ({fmt_pct(max_adj_row['조정비율'])})")

        # 차트
        if chart_kind == "막대+선":
            fig = make_subplots(specs=[[{"secondary_y": True}]])
            fig.add_trace(go.Bar(x=bdf['월'], y=bdf['청구액'],  name='청구액',    marker_color='#3b82f6', opacity=0.85))
            fig.add_trace(go.Bar(x=bdf['월'], y=bdf['조정액'],  name='조정액',    marker_color='#ef4444', opacity=0.85))
            fig.add_trace(go.Bar(x=bdf['월'], y=bdf['심사결정액'], name='심사결정액', marker_color='#22c55e', opacity=0.85))
            fig.add_trace(go.Scatter(x=bdf['월'], y=bdf['조정비율'], name='조정비율(%)',
                                     mode='lines+markers', line=dict(color='#f59e0b', width=3),
                                     marker=dict(size=8)), secondary_y=True)
            fig.update_layout(barmode='group', height=420, plot_bgcolor='white',
                              legend=dict(orientation='h', y=1.08),
                              yaxis_title='금액 (원)', yaxis2_title='조정비율 (%)')
        elif chart_kind == "누적 막대":
            fig = go.Figure()
            fig.add_trace(go.Bar(x=bdf['월'], y=bdf['심사결정액'], name='심사결정액', marker_color='#22c55e'))
            fig.add_trace(go.Bar(x=bdf['월'], y=bdf['조정액'],     name='조정액',    marker_color='#ef4444'))
            fig.update_layout(barmode='stack', height=420, plot_bgcolor='white',
                              yaxis_title='금액 (원)', legend=dict(orientation='h', y=1.08))
        else:
            fig = go.Figure()
            fig.add_trace(go.Scatter(x=bdf['월'], y=bdf['청구액'],     name='청구액',    fill='tonexty', fillcolor='rgba(59,130,246,0.2)', line=dict(color='#3b82f6')))
            fig.add_trace(go.Scatter(x=bdf['월'], y=bdf['심사결정액'], name='심사결정액', fill='tozeroy',  fillcolor='rgba(34,197,94,0.2)',  line=dict(color='#22c55e')))
            fig.add_trace(go.Scatter(x=bdf['월'], y=bdf['조정액'],     name='조정액',    fill='tozeroy',  fillcolor='rgba(239,68,68,0.2)',  line=dict(color='#ef4444')))
            fig.update_layout(height=420, plot_bgcolor='white', yaxis_title='금액 (원)')

        st.plotly_chart(fig, use_container_width=True)

        # 테이블
        st.markdown("**📋 월별 상세 데이터**")
        display_df = bdf[['월','요양급여총액','건수','청구액','조정액','조정비율','심사결정액']].copy()
        display_df['청구액']    = display_df['청구액'].map(lambda x: f"{x:,.0f}")
        display_df['조정액']    = display_df['조정액'].map(lambda x: f"{x:,.0f}")
        display_df['심사결정액'] = display_df['심사결정액'].map(lambda x: f"{x:,.0f}")
        display_df['요양급여총액']= display_df['요양급여총액'].map(lambda x: f"{x:,.0f}")
        display_df['건수']      = display_df['건수'].map(lambda x: f"{x:,.0f}")
        display_df['조정비율']  = display_df['조정비율'].map(lambda x: f"{x:.2f}%")
        st.dataframe(display_df, use_container_width=True, hide_index=True)

    # 건강보험 vs 의료급여 비교
    st.markdown("---")
    st.markdown("**🔄 건강보험 vs 의료급여 비교 (청구액 기준)**")
    c1, c2 = st.columns(2)
    with c1:
        hb_all = combine_billing([hb_out, hb_inp])
        if not hb_all.empty:
            fig2 = px.bar(hb_all, x='월', y=['청구액','조정액'], barmode='group',
                          title='건강보험 청구·조정 현황',
                          color_discrete_map={'청구액':'#3b82f6','조정액':'#ef4444'})
            fig2.update_layout(height=320, plot_bgcolor='white', showlegend=True,
                               legend=dict(orientation='h', y=1.1))
            st.plotly_chart(fig2, use_container_width=True)
    with c2:
        mg_all = combine_billing([mg_out, mg_inp])
        if not mg_all.empty:
            fig3 = px.bar(mg_all, x='월', y=['청구액','조정액'], barmode='group',
                          title='의료급여 청구·조정 현황',
                          color_discrete_map={'청구액':'#8b5cf6','조정액':'#f97316'})
            fig3.update_layout(height=320, plot_bgcolor='white', showlegend=True,
                               legend=dict(orientation='h', y=1.1))
            st.plotly_chart(fig3, use_container_width=True)


# ─────────────────────────────────────────────
# TAB 2: 처방 분석
# ─────────────────────────────────────────────
with tab2:
    st.markdown('<div class="section-header">💊 처방 내역 월별 분석</div>', unsafe_allow_html=True)

    rx_tab = st.radio("진료 구분 선택", ["외래", "입원", "외래+입원 비교"], horizontal=True)

    if rx_tab in ["외래", "입원"]:
        dfs = df_out_sum_f if rx_tab == "외래" else df_inp_sum_f
        title_label = rx_tab

        if dfs.empty:
            st.info("데이터 없음")
        else:
            c1, c2 = st.columns(2)
            with c1:
                fig = px.bar(dfs, x='월', y='금액', title=f'{title_label} 월별 처방 금액',
                             color='금액', color_continuous_scale='Blues')
                fig.update_layout(height=360, plot_bgcolor='white')
                st.plotly_chart(fig, use_container_width=True)
            with c2:
                fig2 = px.bar(dfs, x='월', y='총투', title=f'{title_label} 월별 처방 건수(총투)',
                              color='총투', color_continuous_scale='Greens')
                fig2.update_layout(height=360, plot_bgcolor='white')
                st.plotly_chart(fig2, use_container_width=True)

            # 처방 금액 vs 청구액 비교
            st.markdown("**💡 처방 금액 vs 청구액 비교**")
            bill_ref = all_bill_out if rx_tab == "외래" else all_bill_inp
            if not bill_ref.empty:
                merged = dfs.merge(
                    bill_ref[['월','청구액','심사결정액','조정액']],
                    on='월', how='outer'
                ).fillna(0)
                fig3 = go.Figure()
                fig3.add_trace(go.Scatter(x=merged['월'], y=merged['금액'],    name='처방금액', mode='lines+markers', line=dict(color='#8b5cf6', width=3)))
                fig3.add_trace(go.Scatter(x=merged['월'], y=merged['청구액'],  name='청구액',   mode='lines+markers', line=dict(color='#3b82f6', width=3)))
                fig3.add_trace(go.Scatter(x=merged['월'], y=merged['심사결정액'], name='심사결정액', mode='lines+markers', line=dict(color='#22c55e', width=3, dash='dot')))
                fig3.update_layout(height=380, plot_bgcolor='white', yaxis_title='금액 (원)',
                                   legend=dict(orientation='h', y=1.08))
                st.plotly_chart(fig3, use_container_width=True)

    else:  # 외래+입원 비교
        if df_out_sum_f.empty and df_inp_sum_f.empty:
            st.info("데이터 없음")
        else:
            fig = go.Figure()
            if not df_out_sum_f.empty:
                fig.add_trace(go.Bar(x=df_out_sum_f['월'], y=df_out_sum_f['금액'], name='외래 처방금액', marker_color='#3b82f6'))
            if not df_inp_sum_f.empty:
                fig.add_trace(go.Bar(x=df_inp_sum_f['월'], y=df_inp_sum_f['금액'], name='입원 처방금액', marker_color='#8b5cf6'))
            fig.update_layout(barmode='group', height=400, plot_bgcolor='white',
                              yaxis_title='금액 (원)', title='외래·입원 처방금액 비교',
                              legend=dict(orientation='h', y=1.08))
            st.plotly_chart(fig, use_container_width=True)

            # 구성비 파이
            c1, c2 = st.columns(2)
            out_total = df_out_sum_f['금액'].sum()
            inp_total = df_inp_sum_f['금액'].sum()
            with c1:
                fig_pie = go.Figure(go.Pie(
                    labels=['외래','입원'],
                    values=[out_total, inp_total],
                    hole=0.4,
                    marker_colors=['#3b82f6','#8b5cf6']
                ))
                fig_pie.update_layout(title='처방금액 구성비', height=320)
                st.plotly_chart(fig_pie, use_container_width=True)
            with c2:
                out_cases = df_out_sum_f['총투'].sum()
                inp_cases = df_inp_sum_f['총투'].sum()
                fig_pie2 = go.Figure(go.Pie(
                    labels=['외래','입원'],
                    values=[out_cases, inp_cases],
                    hole=0.4,
                    marker_colors=['#22c55e','#f59e0b']
                ))
                fig_pie2.update_layout(title='처방건수 구성비', height=320)
                st.plotly_chart(fig_pie2, use_container_width=True)


# ─────────────────────────────────────────────
# TAB 3: 의사별 처방
# ─────────────────────────────────────────────
with tab3:
    st.markdown('<div class="section-header">👨‍⚕️ 의사별 처방 현황</div>', unsafe_allow_html=True)

    doc_tab = st.radio("진료 구분", ["외래", "입원"], horizontal=True, key="doc_tab")
    df_doc = df_out_doc if doc_tab == "외래" else df_inp_doc

    if df_doc.empty:
        st.info("데이터 없음")
    else:
        df_doc_f = df_doc[df_doc['월'].isin(sel_months)].copy()
        doctors_list = sorted(df_doc_f['의사'].unique())

        col1, col2 = st.columns([1, 3])
        with col1:
            sel_docs = st.multiselect("의사 선택", doctors_list, default=doctors_list[:6] if len(doctors_list) > 6 else doctors_list)
        with col2:
            doc_metric = st.radio("지표", ["금액", "총투"], horizontal=True, key="doc_metric")

        if sel_docs:
            df_doc_sel = df_doc_f[df_doc_f['의사'].isin(sel_docs)]

            # 월별 의사별 라인 차트
            fig = px.line(
                df_doc_sel, x='월', y=doc_metric, color='의사',
                markers=True, title=f"의사별 월별 처방 {doc_metric}",
                color_discrete_sequence=px.colors.qualitative.Plotly
            )
            fig.update_layout(height=420, plot_bgcolor='white', yaxis_title=doc_metric,
                              legend=dict(orientation='h', y=-0.2))
            st.plotly_chart(fig, use_container_width=True)

            # 의사별 합계 바
            doc_agg = df_doc_sel.groupby('의사')[doc_metric].sum().reset_index().sort_values(doc_metric, ascending=False)
            fig2 = px.bar(doc_agg, x='의사', y=doc_metric, title=f"의사별 총 처방 {doc_metric} 순위",
                          color=doc_metric, color_continuous_scale='Blues',
                          text_auto=True)
            fig2.update_layout(height=360, plot_bgcolor='white')
            st.plotly_chart(fig2, use_container_width=True)

            # 의사별 구성비
            fig3 = px.pie(doc_agg, names='의사', values=doc_metric,
                          title=f"의사별 {doc_metric} 구성비", hole=0.4,
                          color_discrete_sequence=px.colors.qualitative.Set3)
            fig3.update_layout(height=380)
            st.plotly_chart(fig3, use_container_width=True)

            # 테이블
            pivot = df_doc_sel.pivot_table(index='의사', columns='월', values=doc_metric, fill_value=0)
            pivot['합계'] = pivot.sum(axis=1)
            pivot = pivot.sort_values('합계', ascending=False)
            if doc_metric == '금액':
                st.dataframe(pivot.style.format("{:,.0f}"), use_container_width=True)
            else:
                st.dataframe(pivot.style.format("{:,.0f}"), use_container_width=True)


# ─────────────────────────────────────────────
# TAB 4: 항목별 분석
# ─────────────────────────────────────────────
with tab4:
    st.markdown('<div class="section-header">🔍 처방 항목별 분석</div>', unsafe_allow_html=True)

    item_tab = st.radio("진료 구분", ["외래", "입원"], horizontal=True, key="item_tab")
    df_det = df_out_det_f if item_tab == "외래" else df_inp_det_f

    if df_det.empty:
        st.info("데이터 없음")
    else:
        # TOP N 항목
        top_n = st.slider("TOP N 항목", 5, 30, 15)
        item_metric = st.radio("집계 기준", ["금액", "총투"], horizontal=True, key="item_metric")

        item_agg = df_det.groupby('항목')[item_metric].sum().reset_index()
        item_agg = item_agg.sort_values(item_metric, ascending=False).head(top_n)

        c1, c2 = st.columns([2, 1])
        with c1:
            fig = px.bar(
                item_agg.sort_values(item_metric),
                x=item_metric, y='항목', orientation='h',
                title=f"TOP {top_n} 처방 항목 ({item_metric} 기준)",
                color=item_metric, color_continuous_scale='Blues',
                text_auto=True
            )
            fig.update_layout(height=max(350, top_n * 28), plot_bgcolor='white')
            st.plotly_chart(fig, use_container_width=True)
        with c2:
            fig2 = px.pie(item_agg, names='항목', values=item_metric,
                          title="항목 구성비", hole=0.4,
                          color_discrete_sequence=px.colors.qualitative.Set2)
            fig2.update_traces(textposition='inside', textinfo='percent')
            fig2.update_layout(height=400, showlegend=True,
                               legend=dict(orientation='v', font_size=11))
            st.plotly_chart(fig2, use_container_width=True)

        # 선택 항목 월별 추이
        st.markdown("---")
        all_items = sorted(df_det['항목'].unique())
        default_items = item_agg['항목'].tolist()[:5]
        sel_items = st.multiselect("항목 선택 (월별 추이)", all_items, default=default_items)

        if sel_items:
            df_items_sel = df_det[df_det['항목'].isin(sel_items)]
            fig3 = px.line(df_items_sel, x='월', y=item_metric, color='항목',
                           markers=True, title=f"선택 항목 월별 {item_metric} 추이",
                           color_discrete_sequence=px.colors.qualitative.Plotly)
            fig3.update_layout(height=380, plot_bgcolor='white',
                               legend=dict(orientation='h', y=-0.2))
            st.plotly_chart(fig3, use_container_width=True)

        # 히트맵
        st.markdown("---")
        st.markdown("**🗺️ 항목×월 히트맵**")
        top_items_hm = item_agg['항목'].tolist()
        df_hm = df_det[df_det['항목'].isin(top_items_hm)]
        pivot_hm = df_hm.pivot_table(index='항목', columns='월', values=item_metric, fill_value=0)
        # 월 정렬
        ordered_cols = [m for m in MONTHS if m in pivot_hm.columns]
        pivot_hm = pivot_hm[ordered_cols]

        fig_hm = px.imshow(pivot_hm, aspect='auto', color_continuous_scale='Blues',
                           title=f"항목별 월별 {item_metric} 히트맵")
        fig_hm.update_layout(height=max(350, top_n * 26))
        st.plotly_chart(fig_hm, use_container_width=True)


# ─────────────────────────────────────────────
# TAB 5: 삭감 사유 분석
# ─────────────────────────────────────────────
with tab5:
    st.markdown('<div class="section-header">⚠️ 조정(삭감) 사유 분석</div>', unsafe_allow_html=True)

    syu_sheet = st.radio("보험 구분", ["건강보험", "의료급여"], horizontal=True)
    syu_care  = st.radio("진료 구분", ["외래", "입원"], horizontal=True, key="syu_care")
    syu_key   = ('건강보험' if syu_sheet == "건강보험" else '의료급여',
                 'outpatient' if syu_care == "외래" else 'inpatient')
    df_syu = billing_data.get(syu_key, pd.DataFrame())

    if df_syu.empty:
        st.info("해당 조건의 데이터가 없습니다.")
    else:
        df_syu_f = df_syu[df_syu['월'].isin(sel_months)].copy()

        # 조정액 높은 순 정렬
        df_syu_sorted = df_syu_f.sort_values('조정액', ascending=False)

        c1, c2 = st.columns([1, 2])
        with c1:
            fig = px.bar(df_syu_sorted, x='월', y='조정액',
                         title="월별 조정액",
                         color='조정비율',
                         color_continuous_scale='RdYlGn_r',
                         text_auto=True)
            fig.update_layout(height=380, plot_bgcolor='white')
            st.plotly_chart(fig, use_container_width=True)
        with c2:
            fig2 = go.Figure()
            fig2.add_trace(go.Bar(x=df_syu_sorted['월'], y=df_syu_sorted['청구액'],
                                  name='청구액', marker_color='#3b82f6'))
            fig2.add_trace(go.Bar(x=df_syu_sorted['월'], y=df_syu_sorted['조정액'],
                                  name='조정액', marker_color='#ef4444'))
            fig2.add_trace(go.Scatter(x=df_syu_sorted['월'], y=df_syu_sorted['조정비율'],
                                      name='조정비율(%)', yaxis='y2',
                                      mode='lines+markers+text',
                                      text=[f"{v:.1f}%" for v in df_syu_sorted['조정비율']],
                                      textposition='top center',
                                      line=dict(color='#f59e0b', width=3),
                                      marker=dict(size=10)))
            fig2.update_layout(
                barmode='overlay', height=380, plot_bgcolor='white',
                yaxis=dict(title='금액 (원)'),
                yaxis2=dict(title='조정비율 (%)', overlaying='y', side='right', range=[0, max(df_syu_sorted['조정비율'].max()*1.5, 5)]),
                legend=dict(orientation='h', y=1.1)
            )
            st.plotly_chart(fig2, use_container_width=True)

        # 월별 삭감 사유
        st.markdown("---")
        st.markdown("**📌 월별 주요 조정 사유**")

        for _, row in df_syu_f.sort_values('월번호').iterrows():
            if row['주요조정사유'] and row['주요조정사유'] not in ('None', 'ㅡ', '-', ''):
                with st.expander(
                    f"📅 {row['월']} — 조정액: {row['조정액']:,.0f}원 (조정비율 {row['조정비율']:.2f}%)",
                    expanded=(row['조정비율'] > 2)
                ):
                    reasons = row['주요조정사유'].split('\n')
                    for r in reasons:
                        r = r.strip()
                        if r and r not in ('-', 'ㅡ'):
                            st.markdown(f'<div class="reason-box">• {r}</div>', unsafe_allow_html=True)

                    col_a, col_b, col_c = st.columns(3)
                    col_a.metric("청구액",    f"{row['청구액']:,.0f}원")
                    col_b.metric("조정액",    f"{row['조정액']:,.0f}원")
                    col_c.metric("심사결정액", f"{row['심사결정액']:,.0f}원")

        # 사유 키워드 빈도 분석
        st.markdown("---")
        st.markdown("**🔤 조정 사유 키워드 빈도**")
        all_reasons = ' '.join(df_syu_f['주요조정사유'].fillna('').tolist())
        keywords = ['상병 누락', '초과', '조정', '누락', '식대', '낮병동', '신경차단술', '도수치료',
                    '수액', '주사', '입원료', '처방전', '검사', '상병', '비타민', '약제']
        kw_counts = {kw: all_reasons.count(kw) for kw in keywords if all_reasons.count(kw) > 0}
        if kw_counts:
            kw_df = pd.DataFrame(list(kw_counts.items()), columns=['키워드','빈도']).sort_values('빈도', ascending=False)
            fig_kw = px.bar(kw_df, x='키워드', y='빈도', title="조정 사유 키워드 빈도",
                            color='빈도', color_continuous_scale='Reds', text_auto=True)
            fig_kw.update_layout(height=320, plot_bgcolor='white')
            st.plotly_chart(fig_kw, use_container_width=True)


# ─────────────────────────────────────────────
# TAB 6: 원본 데이터
# ─────────────────────────────────────────────
with tab6:
    st.markdown('<div class="section-header">📋 원본 데이터 조회</div>', unsafe_allow_html=True)

    data_sel = st.selectbox("데이터 선택", [
        "청구/삭감 - 건강보험 외래",
        "청구/삭감 - 건강보험 입원",
        "청구/삭감 - 의료급여 외래",
        "청구/삭감 - 의료급여 입원",
        "처방 - 외래 월별 합계",
        "처방 - 입원 월별 합계",
        "처방 - 외래 항목별",
        "처방 - 입원 항목별",
        "처방 - 외래 의사별",
        "처방 - 입원 의사별",
    ])

    data_map = {
        "청구/삭감 - 건강보험 외래":  billing_data.get(('건강보험','outpatient'), pd.DataFrame()),
        "청구/삭감 - 건강보험 입원":  billing_data.get(('건강보험','inpatient'), pd.DataFrame()),
        "청구/삭감 - 의료급여 외래":  billing_data.get(('의료급여','outpatient'), pd.DataFrame()),
        "청구/삭감 - 의료급여 입원":  billing_data.get(('의료급여','inpatient'), pd.DataFrame()),
        "처방 - 외래 월별 합계":       df_out_sum,
        "처방 - 입원 월별 합계":       df_inp_sum,
        "처방 - 외래 항목별":          df_out_det,
        "처방 - 입원 항목별":          df_inp_det,
        "처방 - 외래 의사별":          df_out_doc,
        "처방 - 입원 의사별":          df_inp_doc,
    }

    show_df = data_map.get(data_sel, pd.DataFrame())
    if show_df.empty:
        st.info("데이터 없음")
    else:
        st.markdown(f"**총 {len(show_df):,}행**")
        st.dataframe(show_df, use_container_width=True, height=500)

        csv = show_df.to_csv(index=False, encoding='utf-8-sig')
        st.download_button(
            label="📥 CSV 다운로드",
            data=csv.encode('utf-8-sig'),
            file_name=f"{data_sel.replace(' ','_').replace('/','_')}.csv",
            mime='text/csv'
        )


# ─────────────────────────────────────────────
# 푸터
# ─────────────────────────────────────────────
st.markdown("---")
st.markdown("""
<div style="text-align:center; color:#9ca3af; font-size:12px; padding:16px 0;">
    🏥 2025년 병원 경영 분석 대시보드 &nbsp;|&nbsp; 건강보험·의료급여 청구/삭감 및 처방 내역 종합 분석
</div>
""", unsafe_allow_html=True)
